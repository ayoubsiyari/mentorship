import csv
import io
import smtplib
import uuid
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import formatdate, make_msgid
from datetime import datetime

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from pydantic import BaseModel, EmailStr
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..db import get_db
from ..deps import require_admin
from ..models import User
from ..schemas import UserPublic
from ..settings import settings
from ..security import hash_password

router = APIRouter(prefix="/api/admin", tags=["admin"])


class SendJournalEmailIn(BaseModel):
    email: EmailStr


class BulkEmailIn(BaseModel):
    emails: list[EmailStr]
    subject: str
    content: str  # HTML content


def _send_journal_access_email(email: str, name: str) -> None:
    """Send journal access instructions email to user."""
    subject = "🎓 دليل الوصول إلى منصة Talaria Journal"
    
    html_body = f"""
    <html dir="rtl">
    <body style="font-family: 'Segoe UI', Tahoma, Arial, sans-serif; padding: 20px; background-color: #1a1a2e; margin: 0;">
        <div style="max-width: 600px; margin: 0 auto; background-color: #0f0f23; border-radius: 15px; padding: 30px; border: 1px solid #3730a3; direction: rtl; text-align: right;">
            <div style="text-align: center; margin-bottom: 25px;">
                <h1 style="color: #ffffff; margin: 0; font-size: 28px;">🎓 Talaria Mentorship</h1>
                <p style="color: #a5b4fc; margin-top: 5px;">دليل الوصول إلى Journal</p>
            </div>
            
            <p style="color: #e0e7ff; font-size: 16px; line-height: 1.8;">
                مرحباً <strong style="color: #ffffff;">{name}</strong>،
            </p>
            
            <p style="color: #c7d2fe; font-size: 15px; line-height: 1.8;">
                يسعدنا إرشادك لكيفية الوصول إلى منصة Journal الخاصة بالمنتورشيب.
            </p>
            
            <div style="background-color: #1e1b4b; border-radius: 10px; padding: 20px; margin: 25px 0; border: 1px solid #3730a3;">
                <h3 style="color: #60a5fa; margin-top: 0; font-size: 16px;">📍 الطريقة الأولى: من الموقع الرئيسي</h3>
                <ol style="color: #c7d2fe; font-size: 14px; line-height: 2; padding-right: 20px; margin: 0;">
                    <li>اذهب إلى الموقع الرئيسي <a href="https://talaria-log.com" style="color: #60a5fa;">talaria-log.com</a></li>
                    <li>اضغط على زر <strong style="color: #fbbf24;">2025 Mentorship Login</strong> في أعلى الصفحة</li>
                    <li>سيتم تحويلك إلى صفحة تسجيل الدخول</li>
                    <li>أدخل بريدك الإلكتروني وكلمة المرور</li>
                </ol>
            </div>
            
            <div style="background-color: #052e16; border-radius: 10px; padding: 20px; margin: 25px 0; border: 1px solid #16a34a;">
                <h3 style="color: #4ade80; margin-top: 0; font-size: 16px;">🔗 الطريقة الثانية: الرابط المباشر</h3>
                <p style="color: #bbf7d0; font-size: 14px; line-height: 1.8; margin: 0 0 15px 0;">
                    يمكنك الوصول مباشرة إلى صفحة تسجيل الدخول عبر الرابط التالي:
                </p>
                <div style="text-align: center;">
                    <a href="https://talaria-log.com/journal/login" style="display: inline-block; background-color: #4ade80; color: #052e16; padding: 12px 25px; border-radius: 8px; font-size: 14px; font-weight: bold; text-decoration: none;">
                        🚀 الدخول إلى Journal
                    </a>
                </div>
                <p style="color: #86efac; font-size: 13px; text-align: center; margin-top: 15px; margin-bottom: 0;">
                    https://talaria-log.com/journal/login
                </p>
            </div>
            
            <div style="background-color: #422006; border-radius: 10px; padding: 15px; margin: 25px 0; border: 1px solid #ca8a04;">
                <p style="color: #fef3c7; font-size: 13px; line-height: 1.8; margin: 0;">
                    💡 <strong>نصيحة:</strong> احفظ الرابط المباشر في متصفحك للوصول السريع في المستقبل.
                </p>
            </div>
            
            <p style="color: #94a3b8; font-size: 13px; text-align: center; margin-top: 30px; padding-top: 20px; border-top: 1px solid #3730a3;">
                إذا واجهت أي مشكلة، تواصل معنا على support-center@talaria-log.com
            </p>
            
            <p style="color: #64748b; font-size: 12px; text-align: center; margin-top: 15px;">
                © 2026 Talaria-Log. جميع الحقوق محفوظة.
            </p>
        </div>
    </body>
    </html>
    """

    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = f"Talaria Mentorship <{settings.smtp_from_email or settings.smtp_user}>"
    msg["To"] = email

    msg.attach(MIMEText(html_body, "html"))

    with smtplib.SMTP(settings.smtp_host, settings.smtp_port) as server:
        server.starttls()
        server.login(settings.smtp_user, settings.smtp_password)
        server.send_message(msg)


@router.get("/users")
def list_users(_: User = Depends(require_admin), db: Session = Depends(get_db)):
    users = db.execute(select(User).order_by(User.created_at.desc())).scalars().all()
    return {"users": [UserPublic.model_validate(u, from_attributes=True) for u in users]}


@router.post("/send-journal-email")
def send_journal_email(
    payload: SendJournalEmailIn,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Send journal access instructions email to a specific user."""
    user = db.execute(select(User).where(User.email == payload.email)).scalar_one_or_none()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    _send_journal_access_email(user.email, user.name)
    return {"message": f"Journal access email sent to {user.email}"}


def _send_bulk_email(email: str, subject: str, html_content: str) -> None:
    """Send a custom HTML email to a user. Content is sent as-is without wrapping."""
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = f"Talaria <{settings.smtp_from_email or settings.smtp_user}>"
    msg["To"] = email
    msg["Message-ID"] = make_msgid(domain="talaria-log.com")
    msg["Date"] = formatdate(localtime=True)

    # Send the HTML content as-is without wrapping (templates already have complete structure)
    msg.attach(MIMEText(html_content, "html"))

    with smtplib.SMTP(settings.smtp_host, settings.smtp_port) as server:
        server.starttls()
        server.login(settings.smtp_user, settings.smtp_password)
        server.send_message(msg)


@router.post("/send-bulk-email")
def send_bulk_email(
    payload: BulkEmailIn,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Send a custom email to multiple users at once."""
    sent_count = 0
    errors = []
    
    # Remove duplicates while preserving order
    unique_emails = list(dict.fromkeys([e.lower().strip() for e in payload.emails if e and '@' in e]))
    
    for email in unique_emails:
        try:
            _send_bulk_email(email, payload.subject, payload.content)
            sent_count += 1
        except Exception as e:
            errors.append({"email": email, "error": str(e)})
    
    return {
        "message": f"Email sent to {sent_count} users",
        "total": len(unique_emails),
        "sent": sent_count,
        "failed": len(errors),
        "errors": errors
    }


@router.post("/send-journal-email-all")
def send_journal_email_all(
    _: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Send journal access instructions email to all users with journal access."""
    users = db.execute(select(User).where(User.has_journal_access == True)).scalars().all()
    
    sent_count = 0
    errors = []
    
    for user in users:
        try:
            _send_journal_access_email(user.email, user.name)
            sent_count += 1
        except Exception as e:
            errors.append({"email": user.email, "error": str(e)})
    
    return {
        "message": f"Journal access email sent to {sent_count} users",
        "total_users": len(users),
        "sent": sent_count,
        "errors": errors
    }


@router.post("/import-users")
async def import_users(
    file: UploadFile = File(...),
    user_source: str = Form(default="talaria-prop"),
    has_journal_access: bool = Form(default=False),
    _: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """
    Import users from CSV/XLS file.
    
    Required columns: first_name, last_name, email
    Optional columns: phone, country, password
    
    If password is not provided, a random one will be generated.
    If password looks like a hash (starts with 'scrypt:' or 'pbkdf2:'), it will be used as-is.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")
    
    # Read file content
    content = await file.read()
    
    # Determine file type and parse
    filename_lower = file.filename.lower()
    rows = []
    
    if filename_lower.endswith('.csv'):
        # Parse CSV
        text_content = content.decode('utf-8-sig')  # Handle BOM
        reader = csv.DictReader(io.StringIO(text_content))
        rows = list(reader)
    elif filename_lower.endswith(('.xls', '.xlsx')):
        # Parse Excel - need openpyxl
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1] if cell.value]
            headers = [h.lower().strip().replace(' ', '_') for h in headers]
            
            # Parse rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not any(row):  # Skip empty rows
                    continue
                row_dict = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        row_dict[headers[i]] = value
                rows.append(row_dict)
        except ImportError:
            raise HTTPException(status_code=400, detail="Excel support not available. Please upload CSV.")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Use CSV or XLSX.")
    
    if not rows:
        raise HTTPException(status_code=400, detail="No data found in file")
    
    # Normalize column names
    normalized_rows = []
    for row in rows:
        normalized = {}
        for key, value in row.items():
            if key:
                normalized_key = str(key).lower().strip().replace(' ', '_')
                normalized[normalized_key] = value
        normalized_rows.append(normalized)
    rows = normalized_rows
    
    # Process each row
    created = 0
    skipped = 0
    errors = []
    
    for i, row in enumerate(rows, start=2):  # Start at 2 (row 1 is header)
        try:
            # Get required fields - support both 'name' column or 'first_name'/'last_name'
            name_col = str(row.get('name', '') or '').strip()
            first_name = str(row.get('first_name', '') or '').strip()
            last_name = str(row.get('last_name', '') or '').strip()
            
            # If 'name' column exists, use it; otherwise combine first_name + last_name
            if name_col:
                first_name = name_col
                last_name = ''
            
            email = str(row.get('email', '') or '').strip().lower()
            
            if not email or '@' not in email:
                errors.append({"row": i, "error": "Invalid or missing email"})
                skipped += 1
                continue
            
            # Check if user already exists
            existing = db.execute(select(User).where(User.email == email)).scalar_one_or_none()
            if existing:
                errors.append({"row": i, "email": email, "error": "User already exists"})
                skipped += 1
                continue
            
            # Get optional fields
            phone = str(row.get('phone', '') or '').strip() or None
            country = str(row.get('country', '') or '').strip() or None
            password_raw = str(row.get('password', '') or '').strip()
            
            # Handle password
            if password_raw and (password_raw.startswith('scrypt:') or password_raw.startswith('pbkdf2:')):
                # Already hashed
                password_hash = password_raw
            elif password_raw:
                # Plain text password - hash it
                password_hash = hash_password(password_raw)
            else:
                # Generate random password
                random_pass = str(uuid.uuid4())[:12]
                password_hash = hash_password(random_pass)
            
            # Create user
            name = f"{first_name} {last_name}".strip() or email.split('@')[0]
            
            new_user = User(
                name=name,
                email=email,
                password_hash=password_hash,
                phone=phone,
                country=country,
                role="user",
                is_active=True,
                has_journal_access=has_journal_access,
                email_verified=True,  # Mark as verified since admin is importing
                user_source=user_source,
                created_at=datetime.utcnow()
            )
            db.add(new_user)
            db.commit()  # Commit each user individually to handle duplicates
            created += 1
            
        except Exception as e:
            db.rollback()  # Rollback failed insert
            errors.append({"row": i, "email": email if 'email' in dir() else '', "error": str(e)[:100]})
            skipped += 1
    
    return {
        "message": f"Import completed: {created} users created, {skipped} skipped",
        "created": created,
        "skipped": skipped,
        "total_rows": len(rows),
        "errors": errors[:20]  # Limit errors to first 20
    }


@router.get("/users/by-source/{source}")
def get_users_by_source(
    source: str,
    _: User = Depends(require_admin),
    db: Session = Depends(get_db)
):
    """Get all users from a specific source (e.g., talaria-prop)."""
    users = db.execute(
        select(User).where(User.user_source == source).order_by(User.created_at.desc())
    ).scalars().all()
    return {
        "users": [UserPublic.model_validate(u, from_attributes=True) for u in users],
        "count": len(users)
    }
